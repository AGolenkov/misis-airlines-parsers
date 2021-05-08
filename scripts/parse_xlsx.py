#!/usr/bin/env python

import json
import os

import openpyxl
from tqdm import tqdm

import db


def update_entry(entry: dict) -> dict:
    name = entry["name"].split()

    entry["PassengerLastName"] = name[0] if len(name) >= 1 else ""
    entry["PassengerFirstName"] = name[1] if len(name) >= 2 else ""
    entry["PassengerSecondName"] = name[2] if len(name) >= 3 else ""

    return entry


def parse_sheet(sheet):
    ret = {}
    ret["sequence"] = sheet.cell(1, 8).value
    ret["title"] = sheet.cell(3, 1).value
    ret["name"] = sheet.cell(3, 2).value
    ret["loyalty"] = sheet.cell(3, 6).value
    ret["letter"] = sheet.cell(3, 8).value
    ret["flight"] = sheet.cell(5, 1).value
    # уточнить
    ret["from"] = sheet.cell(5, 4).value
    ret["to"] = sheet.cell(5, 8).value

    ret["gate"] = sheet.cell(7, 2).value
    # уточнить
    ret["from_airport"] = sheet.cell(7, 4).value
    ret["to_airport"] = sheet.cell(7, 8).value

    ret["date"] = sheet.cell(9, 1).value
    ret["time"] = sheet.cell(9, 3).value
    ret["airlines"] = sheet.cell(9, 5).value
    ret["seat"] = sheet.cell(11, 8).value
    ret["PNR"] = sheet.cell(13, 2).value
    ret["e-ticket"] = sheet.cell(13, 5).value

    return ret


def parse_xlsx_file(filepath: str = None) -> list:
    wb = openpyxl.load_workbook(filename=filepath)

    rets = []

    for sheet in wb.worksheets:
        rets.append(parse_sheet(sheet))

    return rets


def parse_xlsx(filepath: str = None) -> None:
    if not filepath:
        with open("config.json", "r") as readfile:
            filepath = json.load(readfile)["data_dir"] + "/" + "YourBoardingPassDotAero/"

    for file in tqdm(os.listdir(filepath)):
        entries = parse_xlsx_file(filepath + file)

        for entry in entries:
            entry = update_entry(entry)
            db.boarding_passes.insert_one(entry)


if __name__ == "__main__":
    parse_xlsx()
